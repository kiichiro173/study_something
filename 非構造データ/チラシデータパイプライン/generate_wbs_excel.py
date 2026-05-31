from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'WBS'
headers = ['ID', '大タスク', '中タスク', '小タスク', '説明', '予定期間', '備考']
ws.append(headers)

rows = [
    [1, 'プロジェクト準備', 'GCPプロジェクト構成', 'GCPプロジェクトの作成/選定', 'GCPプロジェクトを作成し、必要なAPIを有効化する', '', ''],
    [2, 'プロジェクト準備', 'IAM設計', 'サービスアカウント (IAM) の設計・作成', 'BigQuery、GCS、Document AI、Cloud Run 用に必要な権限を定義', '', ''],
    [3, 'プロジェクト準備', 'GCS構成', 'GCSバケットの作成', 'データ受け渡し用バケットを作成', '', ''],
    [4, 'プロジェクト準備', 'データ投入方式検証', 'データアップロード手段の検証', '手動/Cloud Storage API/gsutil などの方法を検証', '', ''],
    [5, 'データ設計', 'スキーマ設計', 'スキーマ（テーブル構造）の設計', 'BigQuery に格納するテーブル構造を定義', '', ''],
    [6, 'データ設計', 'BigQuery構成', 'BigQueryデータセットおよびテーブルの作成', 'データ格納用データセットとテーブルを作成', '', ''],
    [7, '実装', 'メインロジック実装', 'スクリプトの実装', 'PDF画像から構造データを作成するメインロジックを実装', '', ''],
    [8, '実装', 'メインロジック実装', 'Document AI / OCR の精度検証', 'Document AI と他 OCR の精度を比較し、最適手段を決定', '', ''],
    [9, 'デプロイ', 'Cloud Run 構築', 'Cloud Runサービスのデプロイ', '処理を API 形式で提供するため Cloud Run にデプロイ', '', ''],
    [10, '拡張機能', 'Agent Builder', 'Agent Builderの作成', '自動化やインタラクティブな処理を行う Agent を作成', '', ''],
    [11, '検証', 'テスト・チューニング', 'テストとチューニング', 'データが取得・処理できるかを確認し精度を改善', '', ''],
]
for row in rows:
    ws.append(row)
for i, column_cells in enumerate(ws.columns, start=1):
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
    ws.column_dimensions[get_column_letter(i)].width = max_length + 4
wb.save('WBS_for_チラシデータパイプライン.xlsx')
print('Excel file created: WBS_for_チラシデータパイプライン.xlsx')
